# -*- coding: utf-8 -*-
import imaplib
import getpass
import email
from email.mime.text import MIMEText
from email.header import Header
from email.utils import formataddr
from openpyxl import load_workbook
import time

EXCEL_PATH = r"C:\Users\namek\Desktop\Claud codeテスト\名刺一覧.xlsx"
SENDER_NAME = "株式会社乃村工藝社 営業本部"


def load_drafts_from_excel():
    wb = load_workbook(EXCEL_PATH)
    ws = wb["メール下書き"]
    drafts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        no, to_addr, subject, body = row
        if no is not None:
            drafts.append({"no": no, "to": to_addr, "subject": subject, "body": body})
    return drafts


def build_mime_message(from_addr, to_addr, subject, body):
    msg = MIMEText(body, "plain", "utf-8")
    msg["From"] = formataddr((SENDER_NAME, from_addr))
    msg["To"] = to_addr
    msg["Subject"] = Header(subject, "utf-8")
    return msg


def main():
    print("=" * 60)
    print("  乃村工藝社 Gmail下書き一括作成ツール")
    print("=" * 60)
    print()
    print("Gmailアドレスとアプリパスワードを入力してください。")
    print("（アプリパスワード: https://myaccount.google.com/apppasswords）")
    print()

    gmail_address = input("Gmailアドレス: ").strip()
    app_password = getpass.getpass("アプリパスワード（16桁）: ").replace(" ", "").strip()

    print()
    print("Gmailに接続中...")

    try:
        imap = imaplib.IMAP4_SSL("imap.gmail.com", 993)
        imap.login(gmail_address, app_password)
        print("接続成功！")
    except imaplib.IMAP4.error as e:
        print(f"\nエラー: ログインに失敗しました。")
        print("・Gmailアドレスが正しいか確認してください。")
        print("・アプリパスワードを使用しているか確認してください（通常のパスワードは使えません）。")
        print(f"詳細: {e}")
        return

    drafts = load_drafts_from_excel()
    print(f"\nExcelから {len(drafts)} 件のメールを読み込みました。")
    print("Gmailの下書きフォルダに保存中...\n")

    # Gmail の下書きフォルダ名（日本語環境でも "[Gmail]/下書き" の場合あり）
    draft_folders = ['"[Gmail]/Drafts"', '"[Gmail]/下書き"', "Drafts"]
    draft_folder = None

    # 利用可能なフォルダを確認
    _, folders = imap.list()
    for f in folders:
        decoded = f.decode("utf-7") if isinstance(f, bytes) else f
        if "Drafts" in decoded or "下書き" in decoded:
            # フォルダ名を抽出
            parts = decoded.split('"')
            if len(parts) >= 2:
                draft_folder = f'"{parts[-2]}"'
                break
            else:
                draft_folder = decoded.split()[-1]
                break

    if draft_folder is None:
        draft_folder = '"[Gmail]/Drafts"'

    print(f"下書きフォルダ: {draft_folder}")
    print()

    success_count = 0
    for draft in drafts:
        msg = build_mime_message(gmail_address, draft["to"], draft["subject"], draft["body"])
        msg_bytes = msg.as_bytes()

        try:
            result = imap.append(
                draft_folder,
                r"\Draft",
                imaplib.Time2Internaldate(time.time()),
                msg_bytes,
            )
            status = result[0]
            if status == "OK":
                print(f"  [{draft['no']:2d}] ✓ {draft['subject'][:40]}...")
                success_count += 1
            else:
                print(f"  [{draft['no']:2d}] ✗ 保存失敗: {result}")
        except Exception as e:
            print(f"  [{draft['no']:2d}] ✗ エラー: {e}")

    imap.logout()

    print()
    print("=" * 60)
    print(f"  完了: {success_count}/{len(drafts)} 件の下書きを保存しました。")
    print("  Gmailの「下書き」フォルダをご確認ください。")
    print("=" * 60)


if __name__ == "__main__":
    main()
